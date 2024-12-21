from openpyxl import load_workbook

wb = load_workbook('data/stem.xlsx')
ws = wb.active


records = []
for row in ws.iter_rows(min_row = 2):
    school_name = row[0].value
    date = row[3].value.strip()[:-6]
    teacher = row[4].value
    level = row[1].value
    classroom = row[6].value
    mark = row[9].value
    record = {
        'school': school_name,
        'date': date,
        'teacher': teacher,
        'level': level,
        'classroom': classroom,
        'mark': mark,
        'status':'未录入'
    }
    records.append(record)
